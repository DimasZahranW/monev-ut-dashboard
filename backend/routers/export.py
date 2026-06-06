# routers/export.py
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import text
from database import get_db
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io

router = APIRouter(prefix="/export", tags=["Export"])

NAVY  = "0F2167"
TEAL  = "0EA5E9"
WHITE = "FFFFFF"

def thin_border():
    s = Side(style="thin", color="CBD5E1")
    return Border(left=s, right=s, top=s, bottom=s)

def header_style(cell, bg=NAVY):
    cell.font      = Font(name="Arial", bold=True, color=WHITE, size=10)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = thin_border()

def row_style(cell, even=True):
    cell.font      = Font(name="Arial", size=9)
    cell.fill      = PatternFill("solid", start_color="FFFFFF" if even else "F0F4F8")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = thin_border()

def buat_excel(judul, headers, widths, rows, warna_col=None):
    """Helper — buat workbook Excel dengan format standar MONEV"""
    wb = openpyxl.Workbook()
    ws = wb.active

    # ── Baris judul ──
    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    c = ws["A1"]
    c.value     = judul
    c.font      = Font(name="Arial", bold=True, size=12, color=WHITE)
    c.fill      = PatternFill("solid", start_color=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ── Baris header kolom ──
    for col, h in enumerate(headers, 1):
        header_style(ws.cell(row=2, column=col, value=h), bg=TEAL)
    ws.row_dimensions[2].height = 32

    # ── Baris data ──
    for r, row in enumerate(rows, 3):
        even = (r % 2 == 0)
        for col, val in enumerate(row, 1):
            c = ws.cell(row=r, column=col, value=val)
            row_style(c, even=even)
            # Terapkan warna khusus jika ada
            if warna_col and col in warna_col:
                color = warna_col[col].get(str(val))
                if color:
                    c.font = Font(name="Arial", size=9, bold=True, color=color)

    # ── Lebar kolom ──
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{2 + len(rows)}"

    return wb


# ═══════════════════════════════════════════════════════
# 1. EXPORT MAHASISWA
# ═══════════════════════════════════════════════════════
@router.get("/mahasiswa")
def export_mahasiswa(db: Session = Depends(get_db)):
    rows = db.execute(text("""
        SELECT
            a.nim, a.nama_lengkap, a.program_studi, a.upbjj,
            a.jalur_masuk, a.status_pekerjaan,
            k.semester_aktif, k.sks_lulus, k.ipk, k.ips,
            k.status_akademik, k.partisipasi_lms,
            k.prediksi_kelulusan, k.skor_counterfactual
        FROM mahasiswa_admisi a
        JOIN mahasiswa_akademik k ON a.nim = k.nim
        ORDER BY a.nama_lengkap
    """)).fetchall()

    headers = [
        "NIM", "Nama Lengkap", "Program Studi", "UPBJJ",
        "Jalur Masuk", "Status Pekerjaan", "Semester Aktif",
        "SKS Lulus", "IPK", "IPS", "Status Akademik",
        "Partisipasi LMS (%)", "Prediksi Kelulusan", "Skor Counterfactual",
    ]
    widths = [14, 28, 18, 18, 14, 16, 14, 12, 10, 10, 16, 18, 18, 18]

    warna_col = {
        13: {"Lulus": "059669", "Berisiko": "D97706", "DO": "DC2626"}
    }

    wb = buat_excel(
        judul      = "LAPORAN DATA MAHASISWA — MONEV UT 2026",
        headers    = headers,
        widths     = widths,
        rows       = rows,
        warna_col  = warna_col,
    )
    wb.active.title = "Data Mahasiswa"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=laporan_mahasiswa.xlsx"},
    )


# ═══════════════════════════════════════════════════════
# 2. EXPORT DOSEN
# ═══════════════════════════════════════════════════════
@router.get("/dosen")
def export_dosen(db: Session = Depends(get_db)):
    rows = db.execute(text("""
        SELECT
            nidn, nama_lengkap, jabatan_fungsional, unit,
            program_studi, upbjj, h_index, jml_sitasi,
            artikel_scopus, artikel_sinta, hibah_nasional,
            skor_pengajaran, skor_penelitian,
            skor_publikasi, skor_pkm,
            skor_total, label_kinerja, status
        FROM dosen_tutor
        ORDER BY skor_total DESC
    """)).fetchall()

    headers = [
        "NIDN", "Nama Lengkap", "Jabatan", "Unit",
        "Program Studi", "UPBJJ", "H-Index", "Sitasi",
        "Artikel Scopus", "Artikel Sinta", "Hibah Nasional",
        "Skor Pengajaran", "Skor Penelitian",
        "Skor Publikasi", "Skor PkM",
        "Skor Total", "Label Kinerja", "Status",
    ]
    widths = [16, 32, 22, 22, 16, 18, 10, 10, 14, 12, 14, 14, 14, 14, 12, 12, 14, 12]

    warna_col = {
        17: {"Excellent": "059669", "Good": "D97706", "Average": "DC2626"}
    }

    wb = buat_excel(
        judul     = "LAPORAN DATA DOSEN & TUTOR — MONEV UT 2026",
        headers   = headers,
        widths    = widths,
        rows      = rows,
        warna_col = warna_col,
    )
    wb.active.title = "Data Dosen"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=laporan_dosen.xlsx"},
    )


# ═══════════════════════════════════════════════════════
# 3. EXPORT TENDIK
# ═══════════════════════════════════════════════════════
@router.get("/tendik")
def export_tendik(db: Session = Depends(get_db)):
    rows = db.execute(text("""
        SELECT
            nip, nama_lengkap, jabatan, jenis_jabatan,
            unit_kerja, upbjj, status_kepegawaian,
            tingkat_kehadiran, indeks_kepuasan,
            jml_pelatihan, jml_sertifikasi,
            skor_kinerja, label_kinerja, status_aktif
        FROM tendik
        ORDER BY skor_kinerja DESC
    """)).fetchall()

    headers = [
        "NIP", "Nama Lengkap", "Jabatan", "Jenis Jabatan",
        "Unit Kerja", "UPBJJ", "Status Kepegawaian",
        "Kehadiran (%)", "Indeks Kepuasan",
        "Jml Pelatihan", "Jml Sertifikasi",
        "Skor Kinerja", "Label Kinerja", "Status Aktif",
    ]
    widths = [22, 30, 30, 18, 28, 18, 18, 14, 14, 14, 14, 12, 14, 12]

    warna_col = {
        13: {"Excellent": "059669", "Good": "D97706", "Average": "DC2626"}
    }

    wb = buat_excel(
        judul     = "LAPORAN DATA TENDIK — MONEV UT 2026",
        headers   = headers,
        widths    = widths,
        rows      = rows,
        warna_col = warna_col,
    )
    wb.active.title = "Data Tendik"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=laporan_tendik.xlsx"},
    )